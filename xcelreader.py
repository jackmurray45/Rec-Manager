from coordinateClass import Coordinate

import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import datetime
from timeformatter import *
'''
wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        print(col)
        heck = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save(filename = dest_filename)

'''

rec_usage = openpyxl.load_workbook('rec usage.xlsx')


file_name = 'rec usage.xlsx'



def find_bottom(file: 'xlsx file',sheet: 'xcel sheet') -> Coordinate:
    bottom = Coordinate()
    while(file[sheet][bottom.value()].value != None):
        bottom.move_down()
    return bottom



def add_usage_pb(file: 'xlsx file',filename:str, sheet: 'xcel sheet', coord: 'Coordinate', name:str,
                 hall: str, quantity: int, time:datetime.datetime) -> None:
    file[sheet][coord.value()] = give_date(str(time))
    file[sheet][coord.move_right()] = name
    file[sheet][coord.move_right()] = hall
    file[sheet][coord.move_right()] = quantity
    file[sheet][coord.move_right()] = give_time(str(time))
    file[sheet][coord.move_right()] = give_time(str(add_time(time, 30)))
    coord.move_down()
    coord.set_far_left()
    file.save(filename)


def add_usage_vg(file: 'xlsx file', sheet: 'xcel sheet', coord: 'Coordinate', day: 'date', name:str,
                 hall: str, quantity: int, game:str, time:datetime.datetime) -> None:
    file[sheet][coord.value()] = day
    file[sheet][coord.move_right()] = name
    file[sheet][coord.move_right()] = hall
    file[sheet][coord.move_right()] = game
    file[sheet][coord.move_right()] = quantity
    file[sheet][coord.move_right()] = give_time(str(time))
    file[sheet][coord.move_right()] = give_time(str(add_time(time), 30))
    coord.move_down()
    coord.set_far_left()
    rec_usage.save(file_name)

def add_usage_board_and_sports(file: 'xlsx file', sheet: 'xcel sheet', coord: 'Coordinate', day: 'date', name:str,
                               hall: str, game_or_ball: str, time:datetime.datetime) -> None:
    file[sheet][coord.value()] = day
    file[sheet][coord.move_right()] = name
    file[sheet][coord.move_right()] = hall
    file[sheet][coord.move_right()] = game_or_ball
    file[sheet][coord.move_right()] = quantity
    file[sheet][coord.move_right()] = give_time(str(time))
    file[sheet][coord.move_right()] = give_time(str(add_time(time)))
    coord.move_down()
    coord.set_far_left()
    rec_usage.save(file_name)

def add_usage_yoga(file: 'xlsx file', sheet: 'xcel sheet', coord: 'Coordinate', day: 'date',
                   name:str, hall: str,time:datetime.datetime ) -> None:
    file[sheet][coord.value()] = day
    file[sheet][coord.move_right()] = name
    file[sheet][coord.move_right()] = hall
    file[sheet][coord.move_right()] = game_or_ball
    file[sheet][coord.move_right()] = quantity
    file[sheet][coord.move_right()] = give_time(str(time))
    file[sheet][coord.move_right()] = give_time(str(add_time(time)))
    coord.move_down()
    coord.set_far_left()
    rec_usage.save(file_name)


def search_by_name(file: 'xlsx file', sheet: 'xcel sheet', name: str) -> [str]:
    current_sheet = file[sheet]
    coord = Coordinate()
    result = []
    while(current_sheet[coord.value()].value != None or coord.value() != 'Z1'):
        if 'NAME' in current_sheet[coord.value()].value.upper():
            break
        coord.move_right()

    if(current_sheet[coord.value()].value == None or coord.value() == 'Z1'):
        return [None]
        

    name_list = name.upper().split(' ')
    while(current_sheet[coord.value()].value != None):
        current_name = current_sheet[coord.value()].value.split(' ')
        if name_list[0].upper() == current_name[0].upper() and name_list[-1].upper() == current_name[-2].upper():
            result.append(current_sheet[coord.value()].value)
        coord.move_down()
    return result

def return_line(file:' xlsx file', sheet: 'xcel sheet', coord: Coordinate) -> dict:
    result = {}
    title_coord = Coordinate()
    while(file[sheet][title_coord.value()].value != None):
        result[file[sheet][title_coord.value()].value] = file[sheet][coord.value()].value
        title_coord.move_right()
        coord.move_right()
    return result
        
            
    
    
    


    
        
        
        
    
    


